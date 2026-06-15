#!/usr/bin/env python3
"""
SUIT PRO - Enterprise Bi-Directional Spreadsheet Synchronizer Daemon
Provides robust, multi-threaded bidirectionality between PostgreSQL/MySQL datasets and highly styled local Microsoft Excel (.xlsx) registers.
"""

import os
import sys
import time
import json
import csv
import logging
import datetime
import subprocess
from concurrent.futures import ThreadPoolExecutor
from typing import Dict, List, Any

# Configure structured corporate logging
logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] [%(levelname)s] [SUIT-PRO-SYNC-SERVICE] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("suitpro_synchronizer.log", encoding="utf-8")
    ]
)

# Automated background pip check/install for Excel components
try:
    import pandas
    import openpyxl
except ImportError:
    logging.info("Excel dependency libraries not found. Initiating dynamic pip installation package payload...")
    try:
        subprocess.run([sys.executable, "-m", "pip", "install", "--quiet", "pandas", "openpyxl"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=40)
    except Exception as e:
        logging.info(f"Dynamic pip install bypassed: {e}")

# Robust dependency loading for enterprise environment compliance
HAS_POSTGRES = False
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    HAS_POSTGRES = True
except ImportError:
    logging.info("PostgreSQL driver 'psycopg2' is absent. Sync will fallback to local storage databases.")

HAS_PANDAS = False
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    logging.info("Pandas is not installed. Spreadsheet parsing will run in CSV fallback mode.")

HAS_OPENPYXL = False
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    logging.info("openpyxl is not installed. Standard CSV/JSON ledger exports are ready and operational.")

# Core Constants & Target Configurations
DB_HOST = os.getenv("POSTGRES_HOST", "localhost")
DB_NAME = os.getenv("POSTGRES_DB", "suitpro_epos")
DB_USER = os.getenv("POSTGRES_USER", "postgres")
DB_PASSWORD = os.getenv("POSTGRES_PASSWORD", "admin")
DB_PORT = os.getenv("POSTGRES_PORT", "5432")

# Local Storage File Registries
LOCAL_JSON_DB = os.path.join(os.getcwd(), "suitpro_products_db.json")
LOCAL_CSV_LEDGER = os.path.join(os.getcwd(), "suitpro_ledger.csv")
LOCAL_EXCEL_OVERRIDE = os.path.join(os.getcwd(), "suitpro_inventory_override.xlsx")
LOCAL_EXCEL_LEDGER = os.path.join(os.getcwd(), "suitpro_ledger.xlsx")
# Multi-client CSV fallbacks for environments without Pandas/openpyxl
LOCAL_CSV_OVERRIDE = os.path.join(os.getcwd(), "suitpro_inventory_override.csv")

class BidirectionalSynchronizer:
    def __init__(self):
        self.last_sync_timestamp = datetime.datetime.now() - datetime.timedelta(days=1)
        self.thread_pool = ThreadPoolExecutor(max_workers=4, thread_name_prefix="SuitProSyncWorker")
        self.verify_file_structures()

    def verify_file_structures(self):
        """Ensures local Excel sheets, JSON, or CSV databases exist with proper structures on boot."""
        # 1. Excel Inventory Override Sheet Setup
        if HAS_PANDAS and HAS_OPENPYXL:
            if not os.path.exists(LOCAL_EXCEL_OVERRIDE):
                logging.info(f"Creating starter Microsoft Excel override template: {LOCAL_EXCEL_OVERRIDE}")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Inventory Inventory"
                
                # Excel columns
                headers = ["barcode_sku", "name", "size", "colour", "cost_price", "selling_price", "stock_qty"]
                ws.append(headers)
                
                # Starter rows referencing actual items
                ws.append(["88001", "Slim-Fit Midnight Navy Wool Suit", "40R", "Midnight Navy", 320.00, 795.00, 15])
                ws.append(["88002", "Charcoal Double-Breasted Tailored Suit", "42L", "Charcoal Grey", 350.00, 850.00, 7])
                ws.append(["88003", "Premium Peak Lapel Tuxedo Set", "38R", "Jet Black", 400.00, 995.00, 6])
                
                # Styling Excel sheet like a high-end corporate dashboard
                gold_fill = PatternFill(start_color="DFB76C", end_color="DFB76C", fill_type="solid")
                font_bold = Font(name="Segoe UI", size=11, bold=True, color="000000")
                for col_idx, col in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = gold_fill
                    cell.font = font_bold
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Enable auto-fit column widths
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                    
                wb.save(LOCAL_EXCEL_OVERRIDE)
        else:
            # Fallback to generating CSV Inventory Override Template
            if not os.path.exists(LOCAL_CSV_OVERRIDE):
                logging.info(f"Creating starter CSV inventory override template: {LOCAL_CSV_OVERRIDE}")
                try:
                    with open(LOCAL_CSV_OVERRIDE, "w", encoding="utf-8", newline="") as f:
                        writer = csv.writer(f)
                        writer.writerow(["barcode_sku", "name", "size", "colour", "cost_price", "selling_price", "stock_qty"])
                        writer.writerow(["88001", "Slim-Fit Midnight Navy Wool Suit", "40R", "Midnight Navy", "320.00", "795.00", "15"])
                        writer.writerow(["88002", "Charcoal Double-Breasted Tailored Suit", "42L", "Charcoal Grey", "350.00", "850.00", "7"])
                        writer.writerow(["88003", "Premium Peak Lapel Tuxedo Set", "38R", "Jet Black", "400.00", "995.00", "6"])
                except Exception as e:
                    logging.info(f"Failed to create starter CSV override template: {e}")

        # 2. Excel Sales Ledger Sheet Setup
        if not os.path.exists(LOCAL_EXCEL_LEDGER) and HAS_PANDAS and HAS_OPENPYXL:
            logging.info(f"Creating starter Microsoft Excel ledger: {LOCAL_EXCEL_LEDGER}")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales Transactions Ledger"
            headers = ["Invoice ID", "Timestamp", "Items Summary", "Subtotal (GBP)", "VAT 20% (GBP)", "Total Paid (GBP)", "Net Margin Profit (GBP)", "Payment Method", "Salesperson"]
            ws.append(headers)
            
            # Dark luxury style for ledger title
            dark_fill = PatternFill(start_color="121216", end_color="121216", fill_type="solid")
            gold_font = Font(name="Segoe UI", size=11, bold=True, color="DFB76C")
            for col_idx, col in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = dark_fill
                cell.font = gold_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            wb.save(LOCAL_EXCEL_LEDGER)

    def get_postgres_connection(self):
        """Spawns an active connection to the PostgreSQL database if configured and driver is active."""
        if HAS_POSTGRES:
            try:
                return psycopg2.connect(
                    host=DB_HOST,
                    database=DB_NAME,
                    user=DB_USER,
                    password=DB_PASSWORD,
                    port=DB_PORT,
                    connect_timeout=3
                )
            except Exception as err:
                logging.debug(f"PostgreSQL connection offline or failed: {err}")
        return None

    def execute_outbound_sync(self):
        """
        Runs OUTBOUND sync via thread pool:
        Queries any new sales transactions from PostgreSQL (or falls back to parsing suitpro_ledger.csv/JSON entries)
        and adds them to the luxury Microsoft Excel (.xlsx) workbook ledger.
        """
        def outbound_task():
            conn = self.get_postgres_connection()
            new_records = []
            
            if conn:
                try:
                    cur = conn.cursor(cursor_factory=RealDictCursor)
                    query = """
                        SELECT invoice_id, timestamp, items_summary, subtotal, vat_amount, total_due, net_profit, payment_method, salesperson
                        FROM sales_transactions 
                        WHERE timestamp > %s 
                        ORDER BY timestamp ASC
                    """
                    cur.execute(query, (self.last_sync_timestamp,))
                    new_records = cur.fetchall()
                    conn.close()
                except Exception as err:
                    logging.debug(f"Outbound PostgreSQL Query offline: {err}")
            else:
                # No live database connection, fallback to reading suitpro_ledger.csv for changes
                if os.path.exists(LOCAL_CSV_LEDGER):
                    try:
                        with open(LOCAL_CSV_LEDGER, "r", encoding="utf-8") as f:
                            reader = csv.DictReader(f)
                            for r in reader:
                                # Convert parsing formats
                                r_time = datetime.datetime.fromisoformat(r["Timestamp"].replace("Z", "+00:00"))
                                if r_time > self.last_sync_timestamp:
                                    new_records.append({
                                        "invoice_id": r.get("Invoice ID"),
                                        "timestamp": r.get("Timestamp"),
                                        "items_summary": r.get("Items Summary", "Direct sale item"),
                                        "subtotal": float(r.get("Subtotal (GBP)", 0.0)),
                                        "vat_amount": float(r.get("VAT amount (GBP)", 0.0)),
                                        "total_due": float(r.get("Total Paid (GBP)", 0.0)),
                                        "net_profit": float(r.get("Net Profit (GBP)", 0.0)),
                                        "payment_method": r.get("Payment Method"),
                                        "salesperson": r.get("Salesperson")
                                    })
                    except Exception as err:
                        logging.debug(f"Outbound CSV fallback parse status: {err}")

            if new_records and HAS_OPENPYXL:
                try:
                    wb = openpyxl.load_workbook(LOCAL_EXCEL_LEDGER)
                    ws = wb.active
                    
                    logging.info(f"Outbound Outbox: Recording {len(new_records)} transactions to {LOCAL_EXCEL_LEDGER}")
                    thin_border = Border(
                        left=Side(style='thin', color='E0E0E0'),
                        right=Side(style='thin', color='E0E0E0'),
                        top=Side(style='thin', color='E0E0E0'),
                        bottom=Side(style='thin', color='E0E0E0')
                    )
                    
                    for r in new_records:
                        row_vals = [
                            r["invoice_id"],
                            r["timestamp"],
                            r["items_summary"],
                            r["subtotal"],
                            r["vat_amount"],
                            r["total_due"],
                            r["net_profit"],
                            r["payment_method"],
                            r["salesperson"]
                        ]
                        ws.append(row_vals)
                        
                        # Style the newly added data rows beautifully
                        row_idx = ws.max_row
                        for col_idx in range(1, len(row_vals) + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.font = Font(name="Segoe UI", size=10)
                            cell.border = thin_border
                            if col_idx in [4, 5, 6, 7]: # Number formatting for financial rows
                                cell.number_format = '£#,##0.00'
                                
                    wb.save(LOCAL_EXCEL_LEDGER)
                    logging.info("Outbound spreadsheet write completed cleanly.")
                except Exception as err:
                    logging.debug(f"Outbound Excel update failed: {err}")

        # Dispatch via thread pool non-blockingly
        self.thread_pool.submit(outbound_task)

    def execute_inbound_polling_loop(self):
        """
        Inbound Polling: Parses local Excel overrides sheet (LOCAL_EXCEL_OVERRIDE) or falls back to CSV.
        If stock quantities or unit pricing cells have been modified, it cascades the update down
        to PostgreSQL (if online) and concurrently updates the suitpro_products_db.json file
        to reflect coordinates instantly to the live node.js web client.
        """
        logging.info("Inbound Polling Pipeline running...")
        
        records = []
        if HAS_PANDAS:
            if os.path.exists(LOCAL_EXCEL_OVERRIDE):
                try:
                    df = pd.read_excel(LOCAL_EXCEL_OVERRIDE)
                    records = df.to_dict(orient="records")
                except Exception as err:
                    logging.info(f"Excel read skipped or failed: {err}")
        else:
            if os.path.exists(LOCAL_CSV_OVERRIDE):
                try:
                    with open(LOCAL_CSV_OVERRIDE, "r", encoding="utf-8") as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            records.append(row)
                except Exception as err:
                    logging.info(f"CSV fallback read skipped or failed: {err}")
                    
        if not records:
            return

        try:
            updates_applied = 0
            
            # 1. Read existing local products state representation JSON if present
            local_products = []
            if os.path.exists(LOCAL_JSON_DB):
                try:
                    with open(LOCAL_JSON_DB, "r", encoding="utf-8") as f:
                        local_products = json.load(f)
                except Exception as err:
                    logging.info(f"Failed to load products database file: {err}")

            conn = self.get_postgres_connection()
            postgres_cur = conn.cursor() if conn else None

            for rec in records:
                barcode = str(rec.get("barcode_sku", "")).strip()
                if not barcode:
                    continue
                
                # Coerce data types to safety borders
                try:
                    new_price = float(rec.get("selling_price", 0.0))
                    new_cost = float(rec.get("cost_price", 0.0))
                    new_stock = int(rec.get("stock_qty", 0))
                    item_name = str(rec.get("name", ""))
                except (ValueError, TypeError) as val_err:
                    logging.info(f"Inbound Sync formatting error on barcode {barcode}: {val_err}")
                    continue

                # A. Update PostgreSQL database
                if postgres_cur:
                    try:
                        postgres_cur.execute(
                            "UPDATE products SET selling_price = %s, cost_price = %s, stock_qty = %s WHERE barcode_sku = %s",
                            (new_price, new_cost, new_stock, barcode)
                        )
                        if postgres_cur.rowcount > 0:
                            updates_applied += 1
                            logging.info(f"Database Synced: SKU [{barcode}] price set to £{new_price}, stock to {new_stock} units.")
                    except Exception as err:
                        logging.debug(f"PostgreSQL update statement failed for SKU {barcode}: {err}")

                # B. Dynamic Fallback: Synchronize active listing state inside local suitpro_products_db.json
                for p in local_products:
                    if str(p.get("barcode")).strip() == barcode:
                        # Detect modifications first
                        if p.get("sellingPrice") != new_price or p.get("stock") != new_stock or p.get("costPrice") != new_cost:
                            p["sellingPrice"] = new_price
                            p["costPrice"] = new_cost
                            p["stock"] = new_stock
                            updates_applied += 1
                            logging.info(f"Dynamic Json sync: SKU [{barcode}] updated inside local JSON index.")

            if postgres_cur and conn:
                conn.commit()
                conn.close()

            # Save modified JSON index back
            if updates_applied > 0 and local_products:
                try:
                    with open(LOCAL_JSON_DB, "w", encoding="utf-8") as f:
                        json.dump(local_products, f, indent=2)
                    logging.info(f"Dynamic JSON index rewrite success. Applied {updates_applied} updates successfully down to platform registers.")
                except Exception as err:
                    logging.info(f"Failed to commit updated json listings: {err}")

        except Exception as err:
            logging.info(f"Inbound spreadsheet sync cycle aborted/skipped: {err}")

    def run_sync_cycle(self):
        """Coordinates one full bidirectional cycle."""
        try:
            self.execute_outbound_sync()
            self.execute_inbound_polling_loop()
            self.last_sync_timestamp = datetime.datetime.now()
        except Exception as err:
            logging.info(f"Critical synchronization pipeline loop status: {err}")

if __name__ == "__main__":
    logging.info("==========================================================================")
    logging.info("SUIT PRO - Enterprise Bidirectional Excel Sync Service online.")
    logging.info("Outbound multi-threaded ThreadPool and Inbound polling registers online.")
    logging.info("==========================================================================")
    
    sync = BidirectionalSynchronizer()
    
    while True:
        try:
            sync.run_sync_cycle()
        except KeyboardInterrupt:
            logging.info("Exiting sync daemon on operator interrupt signal.")
            break
        except Exception as err:
            logging.info(f"Background daemon sync cycle status: {err}")
        time.sleep(30)
